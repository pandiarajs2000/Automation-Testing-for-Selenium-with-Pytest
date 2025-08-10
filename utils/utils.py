import openpyxl
import logging
from openpyxl.styles import Font, Color

# read data from the excel
def read_data(file_path ,sheet_name, row, column):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.cell(row, column).value

# row count from the excel
def row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    return work_sheet.max_row

# write data to the excel sheet
def write_data(file_path,sheet_name, row, column,data,col2=None, value2=None):
    workbook = openpyxl.load_workbook(file_path)
    workbook.active
    work_sheet = workbook[sheet_name]
    work_sheet.cell(row,column).value = data
    if col2 and value2 is not None:
        work_sheet.cell(row=row, column=col2).value=value2
    return workbook.save(file_path)

def read_all_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = workbook[sheet_name]
    all_data = []
    for row_data in range(2, sheet_name.max_row +1):
        first_name = sheet_name.cell(row=row_data, column=1).value
        last_name = sheet_name.cell(row = row_data, column=2).value
        gstin = sheet_name.cell(row=row_data, column=3).value
        email = sheet_name.cell(row=row_data, column=4).value
        phone = sheet_name.cell(row=row_data, column=5).value
        p_date = sheet_name.cell(row=row_data, column=6).value
        p_time = sheet_name.cell(row=row_data, column=7).value
        
        row_values = [first_name,last_name,gstin,email,phone,p_date,p_time]
        all_data.append(row_values)
    return all_data